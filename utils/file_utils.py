# -*- coding: utf-8 -*-
"""
文件操作工具模块
提供扫描目录、列出Excel文件和Sheet名称等功能
"""

import os
from openpyxl import load_workbook


def list_excel_files(directory="."):
    """
    列出指定目录下的所有xlsx文件
    :param directory: 目标目录路径
    :return: xlsx文件名列表
    """
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    return files


def list_all_sheets(file):
    """
    获取指定Excel文件的所有sheet名称
    :param file: Excel文件路径
    :return: sheet名称列表
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets