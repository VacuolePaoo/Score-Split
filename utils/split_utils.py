# -*- coding: utf-8 -*-

import os
import threading
import psutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock



def process_single_file(args):
    """处理单个文件的函数，用于多线程处理"""
    (file, working_dir, sheet_index, header_row, class_col, 
     student_id_col, ignore_class_col, subject) = args
    
    full_file_path = os.path.join(working_dir, file)
    
    # 使用只读模式打开工作簿以提高性能
    wb = load_workbook(full_file_path, read_only=True, data_only=True)
    
    sheets = wb.sheetnames
    if sheet_index >= len(sheets):
        wb.close()
        return None, f"文件 {file} 没有足够多的sheet"
    
    ws = wb[sheets[sheet_index]]
    
    # 提取表头
    header_data = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    
    # 处理表头（根据需要忽略学号列或班级列）
    if student_id_col is not None or ignore_class_col:
        modified_header = []
        for i, cell in enumerate(header_data):
            # 如果指定了学号列且当前列是学号列，则跳过
            if student_id_col is not None and i + 1 == student_id_col:
                continue
            # 如果需要忽略班级列且当前列是班级列，则跳过
            if ignore_class_col and i + 1 == class_col:
                continue
            modified_header.append(cell)
        subject_header = modified_header
    else:
        subject_header = header_data

    # 提取数据
    file_class_data = {}
    row_count = 0
    
    # 使用values_only=True以提高性能
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        if not row or not row[class_col - 1]:
            continue
            
        class_name = str(row[class_col - 1])
        
        # 如果指定了学号列或需要忽略班级列，则从数据行中移除相应列
        if student_id_col is not None or ignore_class_col:
            modified_row = []
            for i, cell in enumerate(row):
                # 如果指定了学号列且当前列是学号列，则跳过
                if student_id_col is not None and i + 1 == student_id_col:
                    continue
                # 如果需要忽略班级列且当前列是班级列，则跳过
                if ignore_class_col and i + 1 == class_col:
                    continue
                modified_row.append(cell)
            row_data = tuple(modified_row)
        else:
            row_data = row
            
        # 将数据添加到对应班级
        if class_name not in file_class_data:
            file_class_data[class_name] = {}
        if subject not in file_class_data[class_name]:
            file_class_data[class_name][subject] = []
            
        file_class_data[class_name][subject].append(row_data)
        row_count += 1
    
    wb.close()
    
    return (file_class_data, subject_header, row_count), None


def split_and_save(selected_files, sheet_index, sheet_name, header_row, class_col, working_dir=".", student_id_col=None, ignore_class_col=False):
    output_dir = os.path.join(working_dir, "拆分")
    os.makedirs(output_dir, exist_ok=True)

    class_data = {}
    subject_headers = {}
    
    stats = {
        "processed_files": 0,
        "total_files": len(selected_files),
        "generated_classes": 0,
        "total_rows": 0,
        "skipped_files": 0
    }
    
    # 使用线程锁保护共享数据
    class_data_lock = Lock()
    
    total_files = len(selected_files)
    print(f"开始处理 {total_files} 个文件...")
    
    # 使用线程池处理文件以提高性能
    # 使用所有逻辑核心来处理文件，提高处理速度
    max_workers = psutil.cpu_count(logical=True) or 1
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 准备任务参数
        tasks = []
        for file in selected_files:
            subject = os.path.splitext(file)[0]
            tasks.append((
                file, working_dir, sheet_index, header_row, class_col,
                student_id_col, ignore_class_col, subject
            ))
        
        # 提交所有任务
        future_to_file = {executor.submit(process_single_file, task): task[0] for task in tasks}
        
        # 处理完成的任务
        for future in as_completed(future_to_file):
            file = future_to_file[future]
            try:
                result, error = future.result()
                if error:
                    print(f"\n{error}，跳过该文件")
                    stats["skipped_files"] += 1
                else:
                    file_class_data, subject_header, row_count = result
                    
                    # 线程安全地更新共享数据
                    with class_data_lock:
                        # 合并班级数据
                        for class_name, subjects in file_class_data.items():
                            if class_name not in class_data:
                                class_data[class_name] = {}
                            for subject, rows in subjects.items():
                                if subject not in class_data[class_name]:
                                    class_data[class_name][subject] = []
                                class_data[class_name][subject].extend(rows)
                        
                        # 保存表头（假设所有同名学科的表头相同）
                        subject_headers[subject] = subject_header
                        
                        stats["processed_files"] += 1
                        stats["total_rows"] += row_count
                        
            except Exception as e:
                print(f"\n处理文件 {file} 时出错: {e}，跳过该文件")
                stats["skipped_files"] += 1
    
    print("\n数据提取完成，正在生成班级文件...")

    # 按班级排序
    sorted_classes = sorted(class_data.keys(), key=lambda x: int(x) if x.isdigit() else x)
    stats["generated_classes"] = len(sorted_classes)
    
    # 获取当前日期用于制表日期
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # 保存每个班的文件
    total_classes = len(sorted_classes)
    for idx, cls in enumerate(sorted_classes, 1):
        
        subjects = class_data[cls]
        out_file = os.path.join(output_dir, f"{cls}.xlsx")
        # 使用write_only模式提高写入性能
        out_wb = Workbook(write_only=True)
        
        # 按照指定顺序创建sheet
        subject_order = ["语文", "数学", "外语", "物理", "化学", "生物", "历史", "地理", "政治"]
        ordered_subjects = [subject for subject in subject_order if subject in subjects]
        # 添加其他学科
        ordered_subjects.extend([subject for subject in subjects if subject not in ordered_subjects])
        
        # 创建sheet并写入数据
        for subject in ordered_subjects:
            if subject in subjects:  # 确保学科存在
                rows = subjects[subject]
                ws = out_wb.create_sheet(title=subject)
                # 添加标题行，分别放在四个单元格中
                title_row = [f"{subject}", f"{current_date}"]
                # 根据表头长度调整标题行的长度
                if subject in subject_headers and len(subject_headers[subject]) > len(title_row):
                    title_row.extend([""] * (len(subject_headers[subject]) - len(title_row)))
                ws.append(title_row)
                # 使用每个学科自己的表头
                if subject in subject_headers:
                    ws.append(subject_headers[subject])
                for row in rows:
                    # 直接写入元组数据，避免转换为列表的开销
                    ws.append(row)
        
        # 只有当工作簿有工作表时才保存
        if out_wb.worksheets:
            out_wb.save(out_file)
    
    print("\n所有班级文件保存完成!")
    return stats