# -*- coding: utf-8 -*-
"""
Sheet工具模块
处理Excel中sheet的选择功能
"""

import os
from openpyxl import load_workbook
from prompt_toolkit import prompt
from prompt_toolkit.application import get_app, Application
from prompt_toolkit.layout import Layout
from prompt_toolkit.widgets import RadioList, Button, Label, Box
from prompt_toolkit.layout.containers import HSplit, VSplit, Window
from prompt_toolkit.styles import Style


def list_all_sheets(file):
    wb = load_workbook(file, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def choose_sheet(sheets):
    """
    使用单选列表让用户选择学生得分sheet
    """
    # 清屏
    os.system('cls' if os.name == 'nt' else 'clear')
    
    # 创建单选对话框
    values = [(i, sheet) for i, sheet in enumerate(sheets)]
    radio_list = RadioList(values=values)
    radio_list.current_value = 0  # 默认选择第一个
    
    def on_confirm():
        get_app().exit(result=radio_list.current_value)
    
    def on_exit():
        get_app().exit(result="exit")
    
    btn_confirm = Button(text="确认", handler=on_confirm)
    btn_exit = Button(text="退出", handler=on_exit)
    
    # 创建布局
    style = Style.from_dict({
        "button.focused": "reverse",
    })
    
    body = HSplit([
        Label("请选择包含学生得分的sheet:", dont_extend_height=True),
        Window(height=1, char="-"),
        Box(body=radio_list, padding=1),
        Window(height=1, char="-"),
        VSplit([btn_confirm, btn_exit], padding=3),
    ])
    
    application = Application(
        layout=Layout(body, focused_element=radio_list),
        mouse_support=True,
        full_screen=False,
        style=style,
    )
    
    choice = application.run()
    
    if choice == "exit":
        return "exit", "exit"
    elif choice is not None:
        return choice, sheets[choice]
    else:
        return None, None


def ask_sheet_index(sheets):
    print("所有学科的sheet结构相同，以下为第一个文件的sheet列表:")
    for i, sheet in enumerate(sheets):
        print(f"  {i+1}. {sheet}")
    
    while True:
        try:
            index = int(prompt("请输入学生得分sheet的序号: "))
            if 1 <= index <= len(sheets):
                return index-1, sheets[index-1]  # 返回索引和sheet名称
            else:
                print(f"请输入1到{len(sheets)}之间的数字")
        except ValueError:
            print("请输入有效的数字")