# -*- coding: utf-8 -*-
"""
模块名称: exam_splitter
职责范围: 将多个学科成绩单(Excel)按班级拆分，并生成每个班级的成绩汇总表
期望实现计划:
    1. 扫描目录获取所有 xlsx 文件 -> 用户确认
    2. 提取每个文件内的 "学生得分" sheet -> 用户确认
    3. 用户输入表头所在行号
    4. 用户确认班级列所在列号
    5. 按班级排序
    6. 拆分保存到"拆分"文件夹，每个班级一个 Excel，包含各科成绩
已实现功能:
    - 支持交互式选择 (上下键+回车)
    - 自动创建输出文件夹
    - 按班级拆分保存
使用依赖:
    - openpyxl (Excel 处理)
    - prompt_toolkit (交互式菜单)
主要接口:
    - main(): 程序入口
注意事项:
    - 假设所有 Excel 都在当前目录
    - 文件名就是学科名
    - sheet 名需包含 "学生得分"
"""

import os
import warnings
from datetime import datetime
from openpyxl import load_workbook, Workbook
from prompt_toolkit import prompt
from prompt_toolkit.application import Application, get_app
from prompt_toolkit.key_binding import KeyBindings
from prompt_toolkit.layout import Layout
from prompt_toolkit.widgets import CheckboxList, Button, Label, Box, Frame, RadioList, TextArea
from prompt_toolkit.layout.containers import HSplit, VSplit, Window
from prompt_toolkit.styles import Style
# 移除了WordCompleter的导入，因为我们不再需要自动补全功能

# 忽略所有警告
warnings.filterwarnings("ignore")


def choose_working_directory():
    """
    让用户选择工作目录：扫描当前文件夹或手动输入路径
    """
    # 清屏
    os.system('cls' if os.name == 'nt' else 'clear')
    
    # 创建选项对话框
    values = [
        ("current", "扫描本文件夹下工作簿"),
        ("manual", "手动输入文件夹路径")
    ]
    radio_list = RadioList(values=values)
    radio_list.current_value = "current"  # 默认选择当前文件夹
    
    def on_confirm():
        get_app().exit(result=radio_list.current_value)
    
    def on_exit():
        get_app().exit(result="exit")
    
    btn_confirm = Button(text="确认", handler=on_confirm)
    btn_exit = Button(text="退出", handler=on_exit)
    
    # 移除按键绑定，仅保留鼠标支持
    # 创建布局
    style = Style.from_dict({
        "button.focused": "reverse",
    })
    
    body = HSplit([
        Label("请选择工作目录:", dont_extend_height=True),
        Window(height=1, char="-"),
        Box(body=radio_list, padding=1),
        Window(height=1, char="-"),
        VSplit([btn_confirm, btn_exit], padding=3),
    ])
    
    application = Application(
        layout=Layout(body, focused_element=radio_list),
        mouse_support=True,
        full_screen=False,
        style=style,
    )
    
    choice = application.run()
    
    if choice == "exit":
        return "exit", "exit"
    elif not choice:
        return None, None
    
    if choice == "current":
        return ".", "current"
    else:
        # 手动输入路径
        path = prompt("请输入文件夹路径 (支持拖入文件夹获取路径): ").strip().strip('"\'')
        if not os.path.exists(path):
            print(f"路径不存在: {path}")
            return None, None
        if not os.path.isdir(path):
            print(f"路径不是文件夹: {path}")
            return None, None
        return path, "manual"


def list_excel_files(directory="."):
    """
    列出指定目录下的所有xlsx文件
    """
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    return files


def check_output_dir(working_dir="."):
    """
    检查输出目录是否存在文件，并根据情况提供用户选项
    """
    # 清屏
    os.system('cls' if os.name == 'nt' else 'clear')
    
    output_dir = os.path.join(working_dir, "拆分")
    os.makedirs(output_dir, exist_ok=True)
    
    # 检查目录中是否存在任何文件
    existing_files = [f for f in os.listdir(output_dir) if os.path.isfile(os.path.join(output_dir, f))]
    
    # 如果目录为空，直接返回继续执行
    if not existing_files:
        return True
    
    # 如果目录不为空，显示提示并提供选项
    print(f"警告: 输出目录 '{output_dir}' 中已存在以下文件:")
    for f in existing_files[:5]:  # 只显示前5个文件
        print(f"  - {f}")
    if len(existing_files) > 5:
        print(f"  ... 还有 {len(existing_files) - 5} 个文件")
    
    # 创建选项对话框
    values = [
        ("exit", "退出程序，自行处理文件"),
        ("delete", "删除所有现有文件"),
        ("overwrite", "直接覆盖现有文件")
    ]
    radio_list = RadioList(values=values)
    radio_list.current_value = "exit"  # 默认选择退出
    
    def on_confirm():
        get_app().exit(result=radio_list.current_value)
    
    def on_exit():
        get_app().exit(result="exit")
    
    btn_confirm = Button(text="确认", handler=on_confirm)
    btn_exit = Button(text="退出", handler=on_exit)
    
    # 移除按键绑定，仅保留鼠标支持
    # 创建布局
    style = Style.from_dict({
        "button.focused": "reverse",
    })
    
    body = HSplit([
        Label(f"输出目录 '{output_dir}' 中已存在 {len(existing_files)} 个文件:", dont_extend_height=True),
        Window(height=1, char="-"),
        Box(body=radio_list, padding=1),
        Window(height=1, char="-"),
        VSplit([btn_confirm, btn_exit], padding=3),
    ])
    
    application = Application(
        layout=Layout(body, focused_element=radio_list),
        mouse_support=True,
        full_screen=False,
        style=style,
    )
    
    choice = application.run()
    
    if choice == "exit":
        print("用户选择退出程序，请自行处理输出目录中的文件。")
        return False
    elif choice == "delete":
        print("正在删除现有文件...")
        for f in existing_files:
            try:
                os.remove(os.path.join(output_dir, f))
                print(f"  已删除: {f}")
            except Exception as e:
                print(f"  删除 {f} 失败: {e}")
        print("所有现有文件已删除。")
        return True
    elif choice == "overwrite":
        print("用户选择直接覆盖现有文件。")
        return True
    else:
        # 默认退出
        print("操作被取消。")
        return False


def choose_files(files):
    # 清屏
    os.system('cls' if os.name == 'nt' else 'clear')
    
    if not files:
        print("未找到任何xlsx文件")
        return []
    
    # 创建自定义的checkboxlist_dialog
    values = [(f, f) for f in files]
    checkbox = CheckboxList(values=values)
    
    # 创建按钮
    def on_next():
        get_app().exit(result=list(checkbox.current_values or []))
    
    def on_exit():
        get_app().exit(result="exit")
    
    btn_next = Button(text="下一步", handler=on_next)
    btn_select_all = Button(text="全选", handler=lambda: setattr(checkbox, 'current_values', [v for _, v in values]))
    btn_exit = Button(text="退出", handler=on_exit)
    
    # 移除按键绑定，仅保留鼠标支持
    # 创建布局
    style = Style.from_dict({
        "button.focused": "reverse",
    })
    
    body = HSplit([
        Label("请选择要处理的学科 xlsx 文件：", dont_extend_height=True),
        Window(height=1, char="-"),
        Box(body=checkbox, padding=1),
        Window(height=1, char="-"),
        VSplit([btn_select_all, btn_next, btn_exit], padding=3),
    ])
    
    application = Application(
        layout=Layout(body, focused_element=checkbox),
        mouse_support=True,
        full_screen=False,
        style=style,
    )
    
    selected = application.run()
    
    if selected == "exit":
        return "exit"
    return selected or []


def list_all_sheets(file):
    wb = load_workbook(file, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def choose_sheet(sheets):
    """
    使用单选列表让用户选择学生得分sheet
    """
    # 清屏
    os.system('cls' if os.name == 'nt' else 'clear')
    
    # 创建单选对话框
    values = [(i, sheet) for i, sheet in enumerate(sheets)]
    radio_list = RadioList(values=values)
    radio_list.current_value = 0  # 默认选择第一个
    
    def on_confirm():
        get_app().exit(result=radio_list.current_value)
    
    def on_exit():
        get_app().exit(result="exit")
    
    btn_confirm = Button(text="确认", handler=on_confirm)
    btn_exit = Button(text="退出", handler=on_exit)
    
    # 移除按键绑定，仅保留鼠标支持
    # 创建布局
    style = Style.from_dict({
        "button.focused": "reverse",
    })
    
    body = HSplit([
        Label("请选择包含学生得分的sheet:", dont_extend_height=True),
        Window(height=1, char="-"),
        Box(body=radio_list, padding=1),
        Window(height=1, char="-"),
        VSplit([btn_confirm, btn_exit], padding=3),
    ])
    
    application = Application(
        layout=Layout(body, focused_element=radio_list),
        mouse_support=True,
        full_screen=False,
        style=style,
    )
    
    choice = application.run()
    
    if choice == "exit":
        return "exit", "exit"
    elif choice is not None:
        return choice, sheets[choice]
    else:
        return None, None


def ask_sheet_index(sheets):
    print("所有学科的sheet结构相同，以下为第一个文件的sheet列表:")
    for i, sheet in enumerate(sheets):
        print(f"  {i+1}. {sheet}")
    
    while True:
        try:
            index = int(prompt("请输入学生得分sheet的序号: "))  # 移除了completer参数
            if 1 <= index <= len(sheets):
                return index-1, sheets[index-1]  # 返回索引和sheet名称
            else:
                print(f"请输入1到{len(sheets)}之间的数字")
        except ValueError:
            print("请输入有效的数字")


def ask_number(prompt_text):
    while True:
        try:
            user_input = prompt("输入" + prompt_text)
            if user_input.lower() == 'exit':
                return 'exit'
            return int(user_input)
        except ValueError:
            print("输入无效，请输入一个有效的数字！")


def choose_class_column(header_row_data):
    """
    使用单选按钮让用户选择班级所在列
    """
    # 清屏
    os.system('cls' if os.name == 'nt' else 'clear')
    
    # 创建单选列表，最多显示前10列
    values = []
    for i, cell in enumerate(header_row_data[:10]):
        values.append((i+1, f"列{i+1}: {cell}"))
    
    radio_list = RadioList(values=values)
    radio_list.current_value = 1  # 默认选择第一列
    
    def on_confirm():
        get_app().exit(result=radio_list.current_value)
    
    def on_exit():
        get_app().exit(result="exit")
    
    btn_confirm = Button(text="确认", handler=on_confirm)
    btn_exit = Button(text="退出", handler=on_exit)
    
    style = Style.from_dict({
        "button.focused": "reverse",
    })
    
    body = HSplit([
        Label("请选择班级所在列:", dont_extend_height=True),
        Window(height=1, char="-"),
        Box(body=radio_list, padding=1),
        Window(height=1, char="-"),
        VSplit([btn_confirm, btn_exit], padding=3),
    ])
    
    application = Application(
        layout=Layout(body, focused_element=radio_list),
        mouse_support=True,
        full_screen=False,
        style=style,
    )
    
    choice = application.run()
    
    if choice == "exit":
        return "exit"
    elif choice is not None:
        return choice
    else:
        return None


def split_and_save(selected_files, sheet_index, sheet_name, header_row, class_col, working_dir="."):
    output_dir = os.path.join(working_dir, "拆分")
    os.makedirs(output_dir, exist_ok=True)

    class_data = {}  # {class_name: {subject: [rows]}}
    subject_headers = {}  # {subject: header}
    
    total_files = len(selected_files)
    print(f"开始处理 {total_files} 个文件...")
    
    for idx, file in enumerate(selected_files, 1):
        print(f"\r正在处理文件 ({idx}/{total_files}): {file}", end="", flush=True)
        
        # 构造完整的文件路径
        full_file_path = os.path.join(working_dir, file)
        subject = os.path.splitext(file)[0]  # 文件名 = 学科
        wb = load_workbook(full_file_path, read_only=False)
        
        # 使用指定索引获取sheet
        sheets = wb.sheetnames
        if sheet_index < len(sheets):
            ws = wb[sheets[sheet_index]]
        else:
            print(f"\n文件 {file} 没有足够多的sheet，跳过该文件")
            wb.close()
            continue

        # 提取表头（每个学科的表头可能不同）
        header = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        subject_headers[subject] = header

        # 提取数据
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[header_row:]:
            if not row or not row[class_col - 1]:
                continue
            class_name = str(row[class_col - 1])
            if class_name not in class_data:
                class_data[class_name] = {}
            if subject not in class_data[class_name]:
                class_data[class_name][subject] = []
            class_data[class_name][subject].append(row)

        wb.close()

    print("\n数据提取完成，正在生成班级文件...")

    # 按班级排序
    sorted_classes = sorted(class_data.keys(), key=lambda x: int(x) if x.isdigit() else x)
    
    # 获取当前日期用于制表日期
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # 保存每个班的文件
    total_classes = len(sorted_classes)
    for idx, cls in enumerate(sorted_classes, 1):
        print(f"\r正在保存班级文件 ({idx}/{total_classes}): {cls}.xlsx", end="", flush=True)
        subjects = class_data[cls]
        out_file = os.path.join(output_dir, f"{cls}.xlsx")
        out_wb = Workbook()
        out_wb.remove(out_wb.active)  # 删除默认sheet
        for subject, rows in subjects.items():
            ws = out_wb.create_sheet(title=subject)
            # 添加标题行，分别放在四个单元格中
            title_row = [f"{subject}", f"{current_date}"]
            # 根据表头长度调整标题行的长度
            if len(subject_headers[subject]) > len(title_row):
                title_row.extend([""] * (len(subject_headers[subject]) - len(title_row)))
            ws.append(title_row)
            # 使用每个学科自己的表头
            ws.append(subject_headers[subject])
            for row in rows:
                ws.append(row)
        out_wb.save(out_file)
    
    print("\n所有班级文件保存完成!")


def main():
    # 步骤1: 选择工作目录
    working_dir, mode = choose_working_directory()
    if mode == "exit":
        print("程序已退出。")
        return
    if not working_dir:
        print("未选择有效的工作目录，程序退出。")
        return
    
    # 清屏并显示工作目录
    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"工作目录: {os.path.abspath(working_dir)}")
    
    # 步骤2: 检查输出目录
    if not check_output_dir(working_dir):
        return

    # 步骤3: 扫描指定目录，获取所有xlsx文件
    # 清屏并显示文件列表
    os.system('cls' if os.name == 'nt' else 'clear')
    files = list_excel_files(working_dir)
    print("找到以下Excel文件:")
    for f in files:
        print(f"  - {f}")
    
    
    # 步骤4: 让用户选择要处理的文件
    selected = choose_files(files)
    if selected == "exit":
        print("程序已退出。")
        return
    if not selected:
        print("未选择文件，退出。")
        return

    # 步骤5: 获取第一个文件的sheet列表并让用户选择
    # 清屏并显示sheet列表
    os.system('cls' if os.name == 'nt' else 'clear')
    first_file = os.path.join(working_dir, selected[0])
    sheets = list_all_sheets(first_file)
    sheet_index, sheet_name = choose_sheet(sheets)
    
    if sheet_index == "exit":
        print("程序已退出。")
        return
    if sheet_index is None:
        print("未选择sheet，退出。")
        return

    # 步骤6: 询问用户表头所在行数
    # 清屏并询问表头行数
    os.system('cls' if os.name == 'nt' else 'clear')
    header_row = ask_number("表头所在行号: ")
    if header_row == 'exit':
        print("程序已退出。")
        return
    
    # 步骤7: 检索表头行内容并询问用户班级列所在列数
    # 清屏并显示表头内容
    os.system('cls' if os.name == 'nt' else 'clear')
    # 显示第一个文件的表头作为示例
    wb = load_workbook(first_file, read_only=True)
    ws = wb[sheet_name]
    header_row_data = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    wb.close()
    
    print(f"表头行（第{header_row}行）的前10个单元格内容:")
    for i, cell in enumerate(header_row_data[:10]):
        print(f"  列{i+1}: {cell}")
    
    # 使用单选方式选择班级列
    class_col = choose_class_column(header_row_data)
    if class_col == "exit":
        print("程序已退出。")
        return
    if class_col is None:
        print("未选择班级列，退出。")
        return

    # 步骤8-10: 拆分并保存文件
    # 清屏并开始处理文件
    os.system('cls' if os.name == 'nt' else 'clear')
    print("开始拆分文件...")
    split_and_save(selected, sheet_index, sheet_name, header_row, class_col, working_dir)
    print("拆分完成，结果保存在 '拆分' 文件夹中。")


if __name__ == "__main__":
    main()

